import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tkinter import messagebox

def get_input_parameters(entry_threshold, entry_num_sentence, entry_damping):
    try:
        threshold = float(entry_threshold.get())
        if not (0 <= threshold <= 1):
            raise ValueError("Threshold index must be between 0 and 1.")
    except Exception as e:
        messagebox.showerror("Error", f"The value of threshold index is invalid.\n{e}")
        return None

    try:
        _num_sentence_percent = int(entry_num_sentence.get().strip())
        if not (0 <= _num_sentence_percent <= 100):
            raise ValueError("Extracted percentage must be between 0 and 100.")
    except Exception as e:
        messagebox.showerror("Error", f"The value of extracted percentage is invalid.\n{e}")
        return None

    try:
        damping = float(entry_damping.get())
        if not (0 <= damping <= 1):
            raise ValueError("Damping factor d must be between 0 and 1.")
    except Exception as e:
        messagebox.showerror("Error", f"The value of damping factor d is invalid.\n{e}")
        return None

    # Return all values as a tuple
    return threshold, _num_sentence_percent, damping

def log_summary_to_excel(file_name, threshold, damping, num_summary_sentences, num_reference_sentences, match_count, precision, recall, f1_score, percent):
    """
    Ghi log kết quả tóm tắt vào file Excel (.xlsx).
    Tên file Excel = YYYY-MM-DD_percent_threshold_damping.xlsx
    """
    try:
        # Tạo thư mục "document" nếu chưa có
        base_dir = os.path.dirname(os.path.abspath(__file__))
        document_folder = os.path.join(base_dir, "document")
        os.makedirs(document_folder, exist_ok=True)

        # Tạo tên file theo yêu cầu
        now_str = datetime.now().strftime("%Y-%m-%d")
        log_file = os.path.join(
            document_folder,
            f"{now_str}_{percent}_{round(threshold,5)}_{round(damping,5)}.xlsx"
        )

        headers = [
            "Action Time", "File Name", "Threshold", "Damping",
            "Extracted", "Expected", "Correct",
            "Precision", "Recall", "F1-Score"
        ]

        # Nếu file chưa có → tạo mới ghi header
        if not os.path.exists(log_file):
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
        else:
            wb = load_workbook(log_file)
            ws = wb.active

        # Ghi 1 dòng dữ liệu
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = [
            now,
            file_name,
            round(threshold, 5),
            round(damping, 5),
            num_summary_sentences,
            num_reference_sentences,
            match_count,
            round(precision, 5),
            round(recall, 5),
            round(f1_score, 5)
        ]
        ws.append(row)

        # Auto adjust column width
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Save file
        wb.save(log_file)
        return True

    except Exception as e:
        print(f"[ERROR] Failed to write Excel log: {e}")
        return False

def save_summary_to_txt(file_name: str, summary_document: str):
    try:
        # Đường dẫn thư mục đích
        output_dir = r"C:\HUTECH\PROJECT\NLP\document"
        os.makedirs(output_dir, exist_ok=True)

        # Tên file có thêm timestamp để tránh trùng
        base_name = os.path.splitext(os.path.basename(file_name))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_txt_path = os.path.join(output_dir, f"{base_name}_summary_{timestamp}.txt")

        with open(output_txt_path, "w", encoding="utf-8") as f:
            f.write(summary_document)

        print(f"[INFO] Summary written to: {output_txt_path}")
        return output_txt_path

    except Exception as e:
        print(f"[ERROR] Failed to save summary to txt: {e}")
        return None
